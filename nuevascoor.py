import folium
from folium.plugins import FastMarkerCluster
from openpyxl import load_workbook
import requests
import urllib.parse

wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

#locaciones = {}
locaciones = {}

with open("coor3.txt") as f:
    for line in f:
        try:
            key, val = line.split(":")
            locaciones[key] = val
        except ValueError:
            pass

for row in sheet.iter_rows(min_row=6067, max_row=sheet.max_row, values_only=True):
    try:
        if row[2] != None:
            if row[2] not in locaciones:
                url = 'https://nominatim.openstreetmap.org/search/' + urllib.parse.quote(row[2]) +'?format=json'
                response = requests.get(url).json()
                locaciones[row[2]]=[ response[0]["lat"] , response[0]["lon"]]
                with open('coor3.txt', 'a') as f:
                    print(f'{row[2]}:{[ response[0]["lat"] , response[0]["lon"]]}', file=f)
    except IndexError:
        pass
    except UnicodeEncodeError:
        pass
    except TypeError:
        pass

#print(len(locaciones))
#print(locaciones)

