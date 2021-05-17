# Librerías necesarias para leer los datos.
from typing import Tuple
from openpyxl import load_workbook
from openpyxl import Workbook
import time

start = time.time()

# Se suben los datos y se selecciona la hoja a analizar.
wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

# Se leen los títulos para conocer la localización de los elementos a buscar y se despliegan.
tit = []
for row in sheet.iter_rows(min_row=1,max_row=1,values_only=True):
 	tit += row
print(tit)
########## INCLUIR ESTE CODIGO EN LA PRESENTACION ##############

reacciones = { "positivo": ["safe", "treatment", "administration", "administered", "dose", "doses", "health", "healthy",
 "family", "admiration", "courage", "brave", "bravery", "serious", "seriously", "merry", "merrier","Same","paste", "effective"],
"negativo": ["bad", "crime", "cheat", "cheated", "greed", "side", "effects", "corruption", "hurt", "hurts", "hate", "hating",
 "diplomacy", "fake", "stall", "stalled", "war", "ineffective"],
"neutral": ["facts", "fact", "sources", "source", "information", "cases", "case", "deaths", "distribution", "specialist",
 "programme", "inoculation", "inoculating", "needle", "medicine", "symptoms", "available", "update", "schedule", "immunity",
  "authorization", "authorized", "information", "approving", "approved", "manufacture", "manufacturing"]}

vaccines_dict = {"PfizerBioNTech": 0, "AstraZeneca": 0, "SputnikV": 0, "Moderna": 0, "johnsonandjohnson": 0, "Oxford": 0,
 "Novavax": 0, "Sinovac": 0, "Cansino": 0, "Bharat": 0}
vaccines = ["PfizerBioNTech", "AstraZeneca", "SputnikV", "Moderna", "johnsonandjohnson", "Oxford", "Novavax",
 "Sinovac", "Cansino", "Bharat"]

# Búsqueda de tweets mediante palabras clave
Pos = 0
Neg = 0
Neu = 0
total_words = dict()
for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, values_only = True):
    rate = 0
    words = row[10].split()

    if row[11] != None : hashtags = row[11].split("'")

    day = row[9].date()
    # This code line does not really solve a problem, but it makes the classification work efficiently.
    if row[11] != None : hashtags = row[11].split("'")

    """ for t in dates : 
        if t[0] == 0 : dates[t] = dates[1:]
    day = date(",".join(dates)) """

    if row[13] != None : rt = int( row[13] ) / ( today - day ).days
    if rt > max_retweet : max_retweet, best_retweet = rt, row[10]

    if row[14] != None : lk = int( row[14] ) / (today - day).days
    if lk > max_likes : max_likes, best_liked = lk, row[10]

    # Word are clasified and a global rate for the sentence is defined. The lower case convention is used in words and hashtags to avoid the consideration of repeated words with capital lettters present.

    for word in words:
        if word in reacciones["positivo"]:
            rate += 1
        elif word in reacciones["negativo"]:
            rate += -1
        elif word not in total_words:
            total_words[word] = 1
        elif word in total_words:
            total_words[word] += 1
        elif word[0] == "#":
            hashtags.append(word)

    if rate >= 1 : Pos += 1
    elif rate <= -1 : Neg += 1
    else : Neu += 1
    
    vaccines_index = list()
    for hashtag in hashtags :
        if hashtag not in vaccines_index and hashtag in vaccines :
            vaccines_index.append(hashtag)
            vaccines_dict[hashtag] += 1

end = time.time()
print("Positive: ", Pos, "\nNegative: ", Neg, "\nNeutral: ", Neu, "\nvaccines: ", vaccines_dict, "\nTime: ", end - start)