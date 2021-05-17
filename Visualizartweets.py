import folium
from folium.plugins import FastMarkerCluster
from openpyxl import load_workbook
import requests
import urllib.parse

#Crear el mapa como un objeto
m = folium.Map(location=[37.7790262, -122.419906])

#Global tooltip
#tooltip = "Click para más información"

#Procesamiento del excel de Kelloggs

wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

coordenadas = []

locaciones: {

}

i = 0
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
    i+= 1
    if row[2].split(" ")[1]
    try:
        url = 'https://nominatim.openstreetmap.org/search/' + urllib.parse.quote(row[2]) +'?format=json'
        response = requests.get(url).json()
        xy = [ response[0]["lat"] , response[0]["lon"]]
        coordenadas.append(xy)
    except TypeError:
        pass
    except IndexError:
        pass
    print(i)
#Otros marcadores #### ESTE ES EL DE FAST MARKER CLUSTER
folium.plugins.FastMarkerCluster(coordenadas,name='Tiendas de conveniencia').add_to(m)
folium.map.LayerControl().add_to(m)

#Añadir una línea de varios puntos
#folium.vector_layers.PolyLine(coordenadas).add_to(m)

#Generar mapa, ya el html se abre en Chrome y muestra todo visual
#m.save('map.html')
m.save('tweets.html')