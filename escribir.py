import folium
from folium.plugins import FastMarkerCluster
from openpyxl import load_workbook
import requests
import urllib.parse

f = open("coor2.txt","a")

wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

for row in sheet.iter_rows(min_row=5899, max_row=10000, values_only=True):
    try:
        url = 'https://nominatim.openstreetmap.org/search/' + urllib.parse.quote(row[2]) +'?format=json'
        response = requests.get(url).json()
        cor = [row[2], [ response[0]["lat"] , response[0]["lon"]]]
        f.write(f'{cor}\n')
    except TypeError:
        cor = [row[2],"Type Error"]
        f.write(f'{cor}\n')
    except IndexError:
        cor = [row[2], "Index Error"]
        f.write(f'{cor}\n')
    except ConnectionResetError:
        f.close()
    except ConnectionAbortedError:
        f.close()
    except ConnectionRefusedError:
        f.close()
    except ConnectionError:
        f.close()

f.close()