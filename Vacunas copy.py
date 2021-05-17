# Necessary libraries
from typing import Tuple
from openpyxl import load_workbook
import folium
from folium.plugins import FastMarkerCluster
from openpyxl import Workbook
import time
import matplotlib.pyplot as plt
from datetime import date

# This line states the timer of a clock to measure computation time.
start = time.time()

# The datasheet is uploaded to the code so it can be read.
wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

# The first line of the datasheet is read and the locations of the interest variables are deployed.
# tit = []
# for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
# 	tit += row
# #print(tit, "\n")

# After analysing the data, two lists of keywords were constructed. These incude word that may be found in text and also specific hashtags.
reactions = { "positive": ["vaccinated", "VaccinesSaveLives", "safe", "treatment", "administration", "administered", "dose", "doses", "health", "healthy", "family", "admiration", "courage", "brave", "bravery", "serious", "seriously", "merry", "merrier","Same","paste", "effective", "healthcare", "stayhome", "stayathome", "covidiots", "stayathomesavelives", "crushcovid", "sciencematters", "dignity", "health", "hopenews", "scienceovermorons", "dignityhealth", "healthcareworker", "modernavaccine", "takethevaccine", "modern", "modernmedicine", "medicine", "vaccineday", "vaccinocovid", "frontlineworkers", "trustscience", "trust", "science", "facts", "fact", "sources", "source", "information", "cases", "case", "deaths", "distribution", "specialist", "programme", "inoculation", "inoculating", "needle", "medicine", "symptoms", "available", "update", "schedule", "immunity", "authorization", "authorized", "information", "approving", "approved", "manufacture", "manufacturing"],
"negative": ["thrombosis", "bad", "crime", "cheat", "cheated", "greed", "side", "effects", "corruption", "hurt", "hurts", "hate", "hating", "diplomacy", "fake", "stall", "stalled", "war", "ineffective", "hesitation", "whereareallthesickpeople", "sick", "sick people"]}

# Some necessary variables are built.
vaccines = ["pfizer", "astrazeneca", "sputnikv", "moderna", "johnson", "oxford", "novavax", "sinovac", "cansino", "bharat"]
total_vaccines = {"pfizer": 0, "astrazeneca": 0, "sputnikv": 0, "moderna": 0, "johnson": 0, "oxford": 0, "novavax": 0, "sinovac": 0, "cansino": 0, "bharat": 0}

# Search of positive and negative tweets according to the keyword lists.
rates = {"positive": 0, "negative": 0, "neutral": 0}
retweets = {"positive": 0, "negative": 0}
max_retweet, best_retweet, max_likes, best_liked, max_impact, best_impact = 0, "", 0, "", 0, ""
today = date.today()

#The dictionary from which we get the coordinates for the locations
locaciones = {}

with open("locaciones.txt") as f:
    for line in f:
        try:
            key, val = line.split(":")
            locaciones[key] = val
        except ValueError:
            pass

coor = []

for k in locaciones:
    b = locaciones[k].strip("'\n]['").split("', '")
    b = [ float(item) for item in b]
    locaciones[k] = 0
    coor.append(b)

total_words = dict()
for row in sheet.iter_rows(min_row = 2, max_row = sheet.max_row, values_only = True):
    rate = 0
    words = row[10].split()
    day = row[9].date()
    # This code line does not really solve a problem, but it makes the classification work efficiently.
    if row[11] != None : hashtags = row[11].split("'")

    if row[13] != None : rt = int( row[13] )
    if rt > max_retweet : max_retweet, best_retweet = rt, row[10]

    if row[14] != None : lk = int( row[14] )
    if lk > max_likes : max_likes, best_liked = lk, row[10]

    if row[13] != None and row[14] != None : imp = ( rt + lk ) / ( today - day ).days
    if imp > max_impact : max_impact, best_impact = imp, row[10]

    # Word are clasified and a global rate for the sentence is defined. The lower case convention is used in words and hashtags to avoid the consideration of repeated words with capital lettters present.
    for word in words:
        if word.lower() in reactions["positive"] : rate += 1
        elif word.lower() in reactions["negative"] : rate -= 1
    
        if word.lower() not in total_words : total_words[word.lower()] = 1
        elif word.lower() in total_words : total_words[word.lower()] += 1

        # Hashtags present in the text are tracked and appended to the corresponding hashtag list.
        if word[0] == "#" : hashtags.append(word[1:])
    
    # Now, the hashtags are analyazed.
    for hashtag in hashtags :

        # We've decided that this can affect the tweet's rate, as some intentions may not be totally considered.
        if hashtag.lower() in reactions["positive"] : rate += 1
        elif hashtag.lower() in reactions["negative"] : rate -= 1

        # The mention of some vaccines is counted and sorted in a dictionary.
        for vaccine in vaccines :
            if hashtag.lower()[0:len(vaccine)] == vaccine : total_vaccines[vaccine] += 1
    
    # the final rate defines whether the sentence is positive, negative or neutral and these are counted.
    if rate >= 1 : rates["positive"], retweets["positive"] = 1 + rates["positive"], rt + retweets["positive"] 
    elif rate <= -1 : rates["negative"], retweets["negative"] = 1 + rates["negative"], rt + retweets["negative"]
    else : rates["neutral"] += 1

    #Here we count the number of tweets by location
    if row[2] in locaciones:
        locaciones[row[2]] += 1

# Timer is terminated and results are displayed.
end = time.time()

print("Most retweets:\n", best_retweet, "\n\n", 
    "Most liked:\n", best_liked, "\n\n",
    "Greatest impact:\n", best_impact, "\n\n"
)

print("Positive: ", rates["positive"], "\nRetweets: ", retweets["positive"], "\nNegative: ", rates["negative"], "\nRetweets: ",  retweets["negative"], "\nNeutral: ", rates["neutral"], "\nvaccines: ", total_vaccines, "\nTime: ", end - start)

# First try of an histogram (this is still in process).
plt.bar(total_vaccines.keys(), total_vaccines.values(), 1, color = 'b')
plt.xticks(rotation = "vertical")
plt.xlabel("Vacunas")
plt.ylabel("Número de menciones")
plt.title("Menciones de cada vacuna")
plt.show()

plt.bar(rates.keys(), rates.values(), 1, color = 'b')
plt.xlabel("Reacciones")
plt.ylabel("Número de tweets")
plt.title("Reacciones a la vacuna")
plt.show()

locaciones = {k: v for k, v in sorted(locaciones.items(), key=lambda item: item[1])}


top10 = []

for i in range(10):
    top10.append(list(locaciones.items())[len(locaciones)-i-1])

top10nom = []
top10num = []
for i in range(len(top10)):
    top10nom.append(top10[i][0])
    top10num.append(top10[i][1])
plt.bar(top10nom, top10num, 1, color = 'r')
plt.xlabel("Locaciones")
plt.ylabel("Número de tweets")
plt.title("Top 10 locaciones por número de tweets")

plt.show()