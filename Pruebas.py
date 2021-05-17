from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime

wb = load_workbook('vaccination_all_tweets.xlsx')
sheet = wb.worksheets[0]

dates = []

for row in sheet.iter_rows(min_row = 2, max_row = 25, values_only = True):
    #print(row[9].strftime("%m/%d/%Y").split("/"))
    print(row[9].date())

#print(type(dates[0]))
#print(dates[0].year())
#for i in range(len(dates)):
#    print(dates[i].strftime("%m/%d/%Y").split("/"))
#print(dates)